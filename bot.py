from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from openpyxl import Workbook, load_workbook
import os

# ---------------- ВОПРОСЫ ---------------- #

QUESTIONS = [
    "Искривление газопровода?",
    "Наличие кронштейнов?",
    "Наличие футляров?",
    "Высота футляра над полом?",
    "Окраска (спереди, сзади)?",
    "Наличие коррозии?",
    "Раскрытие продольного шва?",
    "Необходимость замены крана?",
    "Наличие ИС?",
    "Наличие САКЗ?",
    "Сборка газопроводов на сгонах, резьбах, уголках?",
    "Наличие проводов на газопроводе?",
    "Наличие труб отопления, водоснабжения и канализации менее 150 мм?",
    "Наличие раковины менее 300 мм?",
    "Наличие электрических розеток менее 300 мм?",
    "Наличие переустройства газопровода?",
    "Замуровка газопровода?",
    "Газопроводы и оборудование в ванных комнатах?",
    "Наличие вентиляционного канала?",
    "Переустройство вентиляционного канала?",
    "Принудительная вентиляция?",
    "Скорость воздуха и температура?",
    "Оценка сварных соединений?",
    "Резьбовые тройники?",
    "Дополнительные сведения?",
    "Напряжение по мультиметру?",
    "Сила тока?",
    "Толщина трубы сверху (УЗК)?",
    "Толщина трубы снизу (УЗК)?",
    "Толщина трубы на опуске (УЗК)?",
    "Течь в футляре?",
    "Течь в кране?",
    "Течь в шланге?",
    "Течь в сварных соединениях?",
    "Течь в теле трубы?",
    "Есть ли нижний футляр?",
    "Есть ли верхний футляр?",
    "Есть ли смежный футляр?",
    "Магнитная дефектоскопия?"
]

NESTED = {
    "Есть ли нижний футляр?": {
        "access": "Есть ли доступ к футляру?(низ)",
        "sub": [
            "Наличие соприкосновения трубы и футляра?(низ)",
            "Наличие гермитизации футляра?(низ)",
            "Газопровод окрашен?(низ)",
            "С газопровода слезла краска?(низ)",
            "Имеется коррозийная пыль?(низ)",
            "Имеется ли очаг коррозии?(низ)",
            "Имеется ли отслоения метала?(низ)"
        ]
    },
    "Есть ли верхний футляр?": {
        "access": "Есть ли доступ к футляру?(верх)",
        "sub": [
            "Наличие соприкосновения трубы и футляра?(верх)",
            "Наличие гермитизации футляра?(верх)",
            "Газопровод окрашен?(верх)",
            "С газопровода слезла краска?(верх)",
            "Имеется коррозийная пыль?(верх)",
            "Имеется ли очаг коррозии?(верх)",
            "Имеется ли отслоения метала?(верх)"
        ]
    },
    "Есть ли смежный футляр?": {
        "access": "Есть ли доступ к футляру?(смежный)",
        "sub": [
            "Наличие соприкосновения трубы и футляра?(смежный)",
            "Наличие гермитизации футляра?(смежный)",
            "Газопровод окрашен?(смежный)",
            "С газопровода слезла краска?(смежный)",
            "Имеется коррозийная пыль?(смежный)",
            "Имеется ли очаг коррозии?(смежный)",
            "Имеется ли отслоения метала?(смежный)"
        ]
    }
}

# полный список вопросов для нумерации и Excel
ALL_QUESTIONS = []
for q in QUESTIONS:
    ALL_QUESTIONS.append(q)
    if q in NESTED:
        ALL_QUESTIONS.append(NESTED[q]["access"])
        ALL_QUESTIONS.extend(NESTED[q]["sub"])

COLUMNS = ["Адрес", "Номер квартиры"] + ALL_QUESTIONS

# ---------------- КЛАВИАТУРА ---------------- #

ANSWER_KB = ReplyKeyboardMarkup(
    [["Да", "Нет"], ["Свой ответ"], ["Изменить ответ"]],
    resize_keyboard=True
)
SAVE_KB = ReplyKeyboardMarkup([["/save"]], resize_keyboard=True)

# ---------------- СЕССИЯ ---------------- #

class Session:
    def __init__(self):
        self.address = None
        self.apartment = None
        self.data = {}
        self.main_index = 0
        self.queue = []
        self.waiting_access_for = None
        self.edit_mode = False
        self.edit_step = None

sessions = {}

# ---------------- ВСПОМОГАТЕЛЬНОЕ ---------------- #

def current_question(s: Session):
    if s.queue:
        return s.queue[0]
    if s.main_index < len(QUESTIONS):
        return QUESTIONS[s.main_index]
    return None

def advance(s):
    if s.queue:
        s.queue.pop(0)
    else:
        s.main_index += 1

# ---------------- КОМАНДЫ ---------------- #

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sessions[update.effective_user.id] = Session()
    await update.message.reply_text("Введите адрес:", reply_markup=ReplyKeyboardRemove())

async def save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    fn = f"obhod_{uid}.xlsx"
    if not os.path.exists(fn):
        await update.message.reply_text("Нет данных")
        return
    await update.message.reply_document(open(fn, "rb"))
    os.remove(fn)
    sessions[uid] = Session()
    await update.message.reply_text("Введите новый адрес:", reply_markup=ReplyKeyboardRemove())

# ---------------- ОСНОВНАЯ ЛОГИКА ---------------- #

async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    text = update.message.text.strip()
    s = sessions.get(uid)
    if not s:
        return

    # --- РЕДАКТИРОВАНИЕ ---
    if s.edit_mode:
        if s.edit_step is None:
            if not text.isdigit():
                await update.message.reply_text("Введите номер вопроса цифрами.")
                return
            idx = int(text) - 1
            if idx < 0 or idx >= len(ALL_QUESTIONS):
                await update.message.reply_text("Такого номера нет, попробуйте ещё раз.")
                return
            question_to_edit = ALL_QUESTIONS[idx]
            if question_to_edit not in s.data:
                await update.message.reply_text(
                    "Этот вопрос ещё не заполнен, выберите другой номер вопроса."
                )
                return
            s.edit_step = question_to_edit
            old = s.data.get(s.edit_step, "—")
            await update.message.reply_text(
                f"{idx+1}. {s.edit_step}\nТекущий ответ: {old}\nВведите новый ответ:",
                reply_markup=ReplyKeyboardRemove()
            )
            return
        else:
            s.data[s.edit_step] = text
            s.edit_mode = False
            s.edit_step = None
            await ask(update, s)
            return

    # --- НАЧАЛО РЕДАКТИРОВАНИЯ ---
    if text == "Изменить ответ":
        answered_questions = [q for q in ALL_QUESTIONS if q in s.data]
        if not answered_questions:
            await update.message.reply_text("Пока нет ответов, которые можно редактировать.")
            return
        msg = "Введите номер вопроса:\n"
        for i, q in enumerate(ALL_QUESTIONS, 1):
            if q in s.data:
                msg += f"{i}. {q}\n"
        s.edit_mode = True
        await update.message.reply_text(msg, reply_markup=ReplyKeyboardRemove())
        return

    # --- АДРЕС / КВАРТИРА ---
    if not s.address:
        s.address = text
        s.data["Адрес"] = text
        await update.message.reply_text("Введите номер квартиры:")
        return

    if not s.apartment:
        # Проверка: первый символ должен быть цифрой
        if not text or not text[0].isdigit():
            await update.message.reply_text("Номер квартиры должен начинаться с цифры. Введите снова:")
            return
        s.apartment = text
        s.data["Номер квартиры"] = text
        await ask(update, s)
        return

    # --- ОСНОВНОЙ ОТВЕТ ---
    q = current_question(s)
    if not q:
        return

    s.data[q] = text

    # Если это главный вопрос с вложенными
    if q in NESTED:
        if text == "Да":
            s.queue = [NESTED[q]["access"]]
            s.waiting_access_for = q
        else:
            advance(s)
        await ask(update, s)
        return

    # Если это вопрос доступа к футляру
    if s.waiting_access_for and q == NESTED[s.waiting_access_for]["access"]:
        if text == "Да":
            s.queue = NESTED[s.waiting_access_for]["sub"].copy()
        else:
            s.queue = []
            advance(s)
        s.waiting_access_for = None
        await ask(update, s)
        return

    # Если это один из под-вопросов
    if s.queue:
        advance(s)
        if not s.queue:
            advance(s)
    else:
        advance(s)

    await ask(update, s)

# ---------------- ВЫВОД ВОПРОСА ---------------- #

async def ask(update, s: Session):
    q = current_question(s)
    if q:
        number = ALL_QUESTIONS.index(q) + 1
        await update.message.reply_text(
            f"{number}. {q}",
            reply_markup=ANSWER_KB
        )
    else:
        save_excel(update.effective_user.id, s.data)
        s.apartment = None
        s.data = {"Адрес": s.address}
        s.main_index = 0
        s.queue = []
        await update.message.reply_text(
            "Квартира сохранена ✅\nВведите следующую или /save",
            reply_markup=SAVE_KB
        )

# ---------------- EXCEL ---------------- #

def save_excel(uid, data):
    fn = f"obhod_{uid}.xlsx"
    if not os.path.exists(fn):
        wb = Workbook()
        ws = wb.active
        ws.append(COLUMNS)
    else:
        wb = load_workbook(fn)
        ws = wb.active
    ws.append([data.get(c, "") for c in COLUMNS])
    wb.save(fn)

# ---------------- ЗАПУСК ---------------- #

def main():
    TOKEN = os.getenv("BOT_TOKEN")
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("save", save))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle))
    print("Бот запущен")
    app.run_polling()

if __name__ == "__main__":
    main()

